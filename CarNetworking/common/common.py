#!/usr/bin/env python
# _*_ coding: utf-8 _*_
# @Time : 2021-09-14 11:05
# @Author : Lionel
# @Version：V 0.1
# @File : common.py
# @desc :==============================================

import time
import random
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
import openpyxl
from selenium.common.exceptions import NoSuchElementException
from CarNetworking.Test_File import readconfig


# 自动读取excel里的测试数据
def read_data(filename, sheetname, origin, destination):
    """

    Args:
        filename:文件名
        sheetname:表名
        origin:起始行数
        destination:终止行数

    Returns:用例数据

    """
    document = openpyxl.load_workbook(filename)
    sheet = document[sheetname]
    if destination == 'maxrow':
        destination = sheet.max_row  # 获取最大行
        # max_column = sheet.max_column  # 获取最大列
    testcases = []  # 创建一个空列表，接受字典
    for i in range(origin, destination + 1):
        case = dict(
            id=sheet.cell(row=i, column=1).value,
            username=sheet.cell(row=i, column=3).value,
            password=sheet.cell(row=i, column=4).value,
            verificationcode=sheet.cell(row=i, column=5).value,
            expected=sheet.cell(row=i, column=6).value)
        testcases.append(case)
    return testcases


# 自动写入数据
def write_result(filename, sheetname, row, column, final_result):
    """

    Args:
        filename: 文件名
        sheetname: 表名
        row: 行数
        column: 列数
        final_result: 测试结果

    Returns:测试结果

    """
    document = openpyxl.load_workbook(filename)
    sheet = document[sheetname]
    sheet.cell(row=row, column=column).value = final_result
    document.save(filename)


class CommonWay:
    # 读取数据
    localRedConfig = readconfig.ReadConfig()

    def __init__(self):
        self.driver = webdriver.Firefox()
        self.driver.maximize_window()

    def click(self, coordinate):
        """

        Args:
            coordinate: 元素定位

        Returns:

        """
        location = self.driver.find_element(By.XPATH, coordinate)
        location.click()

    def sendKey(self, coordinate, data):
        """

        Args:
            coordinate: 元素定位
            data: 插入数据

        Returns:

        """
        location = self.driver.find_element(By.XPATH, coordinate)
        location.send_keys(data)

    def clear(self, coordinate):
        """

        Args:
            coordinate: 元素定位

        Returns:

        """
        location = self.driver.find_element(By.XPATH, coordinate)
        location.clear()

    def openDriver(self):
        """

        Returns:打开网页

        """
        url = self.localRedConfig.get_http('url')
        self.driver.get(url)
        WebDriverWait(self.driver, 10).until(EC.title_is('鑫源车联网TSP管理后台-登陆'))
        return self.driver

    # 登录
    def login(self, filename, sheetname, origin, destination):
        """

        Args:
            filename: 文件名
            sheetname: 表名
            origin: 起始行数
            destination: 终止行数

        Returns:

        """
        cases = read_data(filename, sheetname, origin, destination)
        for case in cases:
            case_id = case['id']
            case_username = case['username']
            case_password = case['password']
            case_verificationcode = case['verificationcode']
            case_expectde = case['expected']
            # 用户名为空，不输入
            if case_username is None:
                pass
            else:
                self.sendKey("//input[@placeholder='请输入登录账号']", case_username)
            # 密码为空，不输入
            if case_password is None:
                pass
            else:
                self.sendKey("//input[@placeholder='请输入密码']", case_password)
            # 验证码为空，不输入
            if case_verificationcode is None:
                pass
            else:
                self.sendKey("//input[@placeholder='请输入验证码']", case_verificationcode)
            self.click("/html/body/div[1]/div/form/div[3]/button")
            time.sleep(2.0)
            try:
                Alert = self.driver.find_element(By.XPATH, '//*[@type="dialog"]')
                if Alert.text == case_expectde:
                    write_result(filename, sheetname, case_id + 1, 7, 'pass')
                else:
                    write_result(filename, sheetname, case_id + 1, 7, 'fail')
            # 判断没有弹窗时，退出循环
            except NoSuchElementException:
                write_result(filename, sheetname, case_id + 1, 7, 'pass')
                break
            self.clear("//input[@placeholder='请输入登录账号']")
            self.clear("//input[@placeholder='请输入密码']")
            self.clear("//input[@placeholder='请输入验证码']")
            WebDriverWait(self.driver, 10).until_not(EC.visibility_of_element_located((By.XPATH, '//*[@type="dialog"]'))
                                                     )

    # ui测试
    def uiTest(self, filename, sheetname, sortdata, moduledata, origin, destination):
        """

        Args:
            filename: 文件名
            sheetname: 表名
            sortdata:类型xpath定位数据
            moduledata:模块xpath定位数据
            origin: 起始行数
            destination: 终止行数

        Returns:

        """
        WebDriverWait(self.driver, 20).until(EC.visibility_of_element_located((By.XPATH, '/html/body/div[1]/div[2]')))
        sort = self.driver.find_element(By.XPATH, sortdata)
        # js注入
        self.driver.execute_script("arguments[0].click();", sort)
        module = self.driver.find_element(By.XPATH, moduledata)
        self.driver.execute_script("arguments[0].click();", module)
        WebDriverWait(self.driver, 20).until(EC.visibility_of_element_located((By.XPATH, '//*[@id="data-search-btn"]')))
        cases = read_data(filename, sheetname, origin, destination)
        for case in cases:
            case_id = case['id']
            Xpath = case['username']
            case_expectde = case['expected']
            if case_expectde is None:
                break
            elif self.driver.find_element(By.XPATH, Xpath).text == case_expectde:
                write_result(filename, sheetname, case_id + 1, 7, 'pass')
            else:
                write_result(filename, sheetname, case_id + 1, 7, 'fail')

    # 筛选测试
    def screen(self, filename, sheetname, origin, destination):
        """

        Args:
            filename: 文件名
            sheetname: 表名
            origin: 起始行数
            destination: 终止行数

        Returns:

        """
        cases = read_data(filename, sheetname, origin, destination)
        for case in cases:
            case_id = case['id']
            case_xpath = case['username']
            case_data = case['password']
            expected = case['expected']
            case_test = case_xpath.split()
            data_test = case_data.split()
            long = len(case_test)
            WebDriverWait(self.driver, 10).until(EC.visibility_of_element_located((By.XPATH,
                                                                                   '//*[@id="data-search-btn"]')))
            for i in range(0, long):
                self.click(case_test[i])
                # res_case = self.driver.find_element(By.XPATH, case_test[i])
                # self.driver.execute_script("arguments[0].click();", res_case)
                try:
                    self.driver.find_element_by_xpath(data_test[i]).click()
                except:
                    send = self.driver.find_element_by_xpath(case_test[i])
                    send.send_keys(data_test[i])
            time.sleep(2.0)
            button = self.driver.find_element(By.XPATH, '//*[@id="data-search-btn"]')
            self.driver.execute_script("arguments[0].click();", button)
            try:
                Result = self.driver.find_element(By.XPATH, '/html/body/div[1]/div[6]/div[2]/div[1]/div/div/div/div[2]/\
                                                            div[2]').text
                assert '无数据' not in Result or Result == expected
                write_result(filename, sheetname, case_id + 1, 7, 'pass')
            except AssertionError:
                write_result(filename, sheetname, case_id + 1, 7, 'fail')
            self.driver.refresh()
            time.sleep(2.0)

    # 添加用户
    def addUser(self, filename, sheetname, origin, destination):
        """

        Args:
            filename: 文件名
            sheetname: 表名
            origin: 起始行数
            destination: 终止行数

        Returns:

        """
        cases = read_data(filename, sheetname, origin, destination)
        self.click('//*[@id="message-create"]')
        WebDriverWait(self.driver, 20).until(EC.visibility_of_element_located((By.XPATH, '//*[@id="cedit"]')))
        for case in cases:
            case_id = case['id']
            case_username = case['username']
            case_password = case['password']
            case_cfpassword = case['verificationcode']
            expected = case['expected']
            # 用户名为空，不输入
            if case_username is None:
                pass
            elif case_username == 'random':
                randomnub = random.randint(13311111111, 13399999999)
                self.sendKey("/html/body/div[3]/div[2]/div/form/div[1]/div/div/input", randomnub)
                self.sendKey("/html/body/div[3]/div[2]/div/form/div[2]/div/div/input", case_password)
                self.sendKey("/html/body/div[3]/div[2]/div/form/div[4]/div/div/input", case_cfpassword)
                self.click('/html/body/div[3]/div[2]/div/form/div[5]/div/button')
                Alert = self.driver.find_element(By.XPATH, '//*[@type="dialog"]')
                if Alert.text == expected:
                    write_result(filename, sheetname, case_id + 1, 7, 'pass')
                else:
                    write_result(filename, sheetname, case_id + 1, 7, 'fail')
                break
            else:
                self.sendKey("/html/body/div[3]/div[2]/div/form/div[1]/div/div/input", case_username)
            # 密码为空，不输入
            if case_password is None:
                pass
            else:
                self.sendKey("/html/body/div[3]/div[2]/div/form/div[2]/div/div/input", case_password)
            # 重复密码为空，不输入
            if case_cfpassword is None:
                pass
            else:
                self.sendKey("/html/body/div[3]/div[2]/div/form/div[4]/div/div/input", case_cfpassword)
            self.click('/html/body/div[3]/div[2]/div/form/div[5]/div/button')
            try:
                Alert = self.driver.find_element(By.XPATH, '/html/body/div[6]/div[2]')
                if Alert.text == expected:
                    write_result(filename, sheetname, case_id + 1, 7, 'pass')
                else:
                    write_result(filename, sheetname, case_id + 1, 7, 'fail')
                self.click('/html/body/div[6]/div[3]/a')
            except:
                break
            self.clear("/html/body/div[3]/div[2]/div/form/div[1]/div/div/input")
            self.clear("/html/body/div[3]/div[2]/div/form/div[2]/div/div/input")
            self.clear("/html/body/div[3]/div[2]/div/form/div[4]/div/div/input")
            time.sleep(2.0)

    # def LockUser(self):
    #     WebDriverWait(self.driver, 10).until_not(EC.visibility_of_element_located((By.XPATH, '//*[@type="dialog"]')))
    #     lock = By.XPATH('/html/body/div[1]/div[6]/div[2]/div[1]/div/div/div/div[2]/div[4]/div[2]/table/tbody/tr[2]/td[2]/div')
    #     lock.click()
    #     try:
    #         assert By.XPATH('/html/body/div[1]/div[6]/div[2]/div[1]/div/div/div/div[2]/div[4]/div[2]/table/tbody/tr[2]/td[2]/div/div').text == '锁定'
    #         print('用户正常锁定')
    #     except AssertionError:
    #         print('用户锁定失败')
    #     lock.click()
    #     try:
    #         assert By.XPATH('/html/body/div[1]/div[6]/div[2]/div[1]/div/div/div/div[2]/div[4]/div[2]/table/tbody/tr[2]/td[2]/div/div').text == '正常'
    #         print('锁定用户正常解锁')
    #     except AssertionError:
    #         print('锁定用户解锁失败')

    # 编辑用户
    def compileUser(self, filename, sheetname, origin, destination):
        """

        Args:
            filename: 文件名
            sheetname: 表名
            origin: 起始行数
            destination: 终止行数

        Returns:

        """
        WebDriverWait(self.driver, 10).until_not(EC.visibility_of_element_located((By.XPATH, '//*[@type="dialog"]')))
        cases = read_data(filename, sheetname, origin, destination)
        self.click('/html/body/div[1]/div[6]/div[2]/div[1]/div/div/div/div[2]/div[4]/div[2]/table/tbody/tr[1]/td[3]/div')
        for case in cases:
            case_id = case['id']
            case_phone = case['username']
            expected = case['expected']
            if case_phone is None:
                pass
            elif case_phone == 'random':
                randomnub = random.randint(18800000000, 18899999999)
                self.sendKey('/html/body/div[4]/div[2]/div/form/div[1]/div[2]/div/input', randomnub)
                self.click('/html/body/div[4]/div[2]/div/form/div[2]/div/button')
                try:
                    Alert = self.driver.find_element(By.XPATH, '//*[@type="dialog"]')
                    if Alert.text == expected:
                        write_result(filename, sheetname, case_id + 1, 7, 'pass')
                    else:
                        write_result(filename, sheetname, case_id + 1, 7, 'fail')
                    self.click('/html/body/div[4]/span[1]/a')
                    break
                except NoSuchElementException:
                    write_result(filename, sheetname, case_id + 1, 7, '未弹窗')
            else:
                self.sendKey('/html/body/div[4]/div[2]/div/form/div[1]/div[2]/div/input', case_phone)
            self.click('/html/body/div[4]/div[2]/div/form/div[2]/div/button')
            try:
                Alert = self.driver.find_element(By.XPATH, '/html/body/div[6]/div[2]')
                if Alert.text == expected:
                    write_result(filename, sheetname, case_id + 1, 7, 'pass')
                else:
                    write_result(filename, sheetname, case_id + 1, 7, 'fail')
                self.click('/html/body/div[6]/div[3]/a')
            except NoSuchElementException:
                write_result(filename, sheetname, case_id + 1, 7, '未弹窗')
            self.clear('/html/body/div[4]/div[2]/div/form/div[1]/div[2]/div/input')

    # 创建消息
    def messageCreate(self, filename, sheetname, origin, destination):
        """

        Args:
            filename: 文件名
            sheetname: 表名
            origin: 起始行数
            destination: 终止行数

        Returns:

        """
        WebDriverWait(self.driver, 20).until(EC.visibility_of_element_located((By.XPATH, '//*[@id="message-create"]')))
        cases = read_data(filename, sheetname, origin, destination)
        self.click('//*[@id="message-create"]')
        WebDriverWait(self.driver, 10).until(EC.visibility_of_element_located((By.XPATH, '//*[@id="btn-add-message"]')))
        for case in cases:
            case_id = case['id']
            case_title = case['username']
            case_group = case['verificationcode']
            case_data = case['password']
            expected = case['expected']
            group = case_group.split()
            if case_title is None:
                pass
            else:
                self.sendKey('/html/body/div[4]/div[2]/div/form/div[1]/div/div/input', case_title)

            if '全体' in group:
                self.click('/html/body/div[4]/div[2]/div/form/div[4]/div/div[1]/div')
            elif '特定群体正常' in group:
                self.click('/html/body/div[4]/div[2]/div/form/div[4]/div/div[2]/div[1]')
                self.click('/html/body/div[4]/div[2]/div/form/div[4]/div/div[2]/div[2]/div[1]/div/div/div/input')
                self.click('/html/body/div[4]/div[2]/div/form/div[4]/div/div[2]/div[2]/div[1]/div/div/dl/dd[2]')
            elif '特定群体异常' in group:
                self.click('/html/body/div[4]/div[2]/div/form/div[4]/div/div[2]/div[1]')
            elif '特定车主正常' in group:
                self.click('/html/body/div[4]/div[2]/div/form/div[4]/div/div[3]/div[1]')
                self.sendKey('/html/body/div[4]/div[2]/div/form/div[4]/div/div[3]/div[2]/div/div/input', '18812341234')
            elif '特定车主异常' in group:
                self.click('/html/body/div[4]/div[2]/div/form/div[4]/div/div[3]/div[1]')
            elif '特定车架号正常' in group:
                self.click('/html/body/div[4]/div[2]/div/form/div[4]/div/div[4]/div[1]')
                self.sendKey('/html/body/div[4]/div[2]/div/form/div[4]/div/div[4]/div[2]/div/div/input',
                             'LM6LKD2F2LX143111')
            elif '特定车架号异常' in group:
                self.click('/html/body/div[4]/div[2]/div/form/div[4]/div/div[4]/div[1]')

            if '即时' in group:
                self.click('/html/body/div[4]/div[2]/div/form/div[2]/div/div/div[1]/div')
            elif '定时' in group:
                self.click('/html/body/div[4]/div[2]/div/form/div[2]/div/div/div[2]/div')
                self.click('//*[@id="publishDate"]')
                self.click('/html/body/div[5]/div[2]/div/span[2]')

            if case_data is None:
                pass
            else:
                # 切换进frame中操作
                self.driver.switch_to.frame(self.driver.find_element_by_tag_name('iframe'))
                # 直接在body中输入
                send = self.driver.find_element_by_tag_name('body')
                send.send_keys(case_data)
                # 退出frame
                self.driver.switch_to.default_content()
            self.click('//*[@id="btn-add-message"]')

            try:
                Alert = self.driver.find_element(By.XPATH, '/html/body/div[6]/div[2]')
                if Alert.text == expected:
                    write_result(filename, sheetname, case_id + 1, 7, 'pass')
                else:
                    write_result(filename, sheetname, case_id + 1, 7, 'fail')
                self.click('/html/body/div[6]/div[3]/a')
            except NoSuchElementException:
                write_result(filename, sheetname, case_id + 1, 7, '未弹窗')
            WebDriverWait(self.driver, 10).until_not(EC.visibility_of_element_located((By.XPATH,
                                                                                       '/html/body/div[6]/div[1]')))
            try:
                self.click('/html/body/div[4]/div[2]/div/form/div[1]/div/div/input')
                self.clear('/html/body/div[4]/div[2]/div/form/div[1]/div/div/input')
                self.driver.switch_to.frame(self.driver.find_element_by_tag_name('iframe'))
                self.driver.find_element_by_tag_name('body').clear()
                self.driver.switch_to.default_content()
            except:
                self.click('//*[@id="message-create"]')
                WebDriverWait(self.driver, 10).until(EC.visibility_of_element_located((By.XPATH,
                                                                                       '//*[@id="btn-add-message"]')))

    # 清除之前的结果
    def cleanResult(self, filename, sheetname):
        """

        Args:
            filename: 文件名
            sheetname: 表名

        Returns:

        """
        document = openpyxl.load_workbook(filename)
        sheet = document[sheetname]
        for i in range(2, sheet.max_row + 1):
            write_result(filename, sheetname, i, 7, "")