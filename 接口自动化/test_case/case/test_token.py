import time
import unittest
import requests
import openpyxl
import ast

class TestToken(unittest.TestCase):

    # 取token
    def token(self):
        self.url = 'http://192.168.7.131:8009/auth/login'
        self.data = {
            "username": "18375761940",
            "password": "78417416f116fd746af8c17099c7713a",
            "clientId": "12345",
            "clientType": 1
        }
        self.header = {
            'Content-Type': 'application/json;charset=utf-8',
            'Host': '192.168.7.131:8009',
            'User-Agent': 'okhttp/3.10.0',
            'Accept-Encoding': 'gzip'
        }
        res = requests.post(url=self.url, json=self.data, headers=self.header)
        res_json = res.json()
        self.res_token = res_json['content']['token']
        return self.res_token

    def postway(self, url, json, header):
        self.res = requests.post(url=url, json=json, headers=header)
        self.res_json = self.res.json()
        return self.res_json

    # 自动读取excel里的测试数据
    def read_data(self, filename, sheetname):
        self.document = openpyxl.load_workbook(filename)
        self.sheet = self.document[sheetname]
        self.max_row = self.sheet.max_row  # 获取最大行
        # max_column = sheet.max_column  # 获取最大列
        testcases = []  # 创建一个空列表，接受字典
        for i in range(2, self.max_row+1):
            case = dict(
                id=self.sheet.cell(row=i, column=1).value,
                url=self.sheet.cell(row=i, column=3).value,
                header=self.sheet.cell(row=i, column=4).value,
                data=self.sheet.cell(row=i, column=5).value,
                expected=self.sheet.cell(row=i, column=6).value)
            testcases.append(case)
        return testcases

    # 自动写入数据
    def write_result(self, filename, sheetname, row, column, final_result):
        self.document = openpyxl.load_workbook(filename)
        self.sheet = self.document[sheetname]
        self.sheet.cell(row=row, column=column).value = final_result
        self.document.save(filename)

    # 自动化测试
    def execute_func(self, filename, sheetname):
        self.caces = self.read_data(filename, sheetname)
        for case in self.caces:
            case_id = case['id']
            case_url = case['url']
            case_data = case['data']
            # case_header = Test_Case['header']
            case_expected = ast.literal_eval(case['expected'])
            real_token = {'Authorization': self.token()}
            result = self.postway(case_url, eval(case_data), real_token)
            case_expected_code = case_expected['code']
            result_code = result['code']
            if case_expected_code == result_code:
                self.write_result(filename, sheetname, case_id + 1, 7, 'pass')
            else:
                self.write_result(filename, sheetname, case_id + 1, 7, 'fail')

    def test_password(self):
        self.execute_func('接口自动化用例.xlsx', 'password')

    def test_star(self):
        self.execute_func('接口自动化用例.xlsx', 'star')

    def test_serve(self):
        self.execute_func('接口自动化用例.xlsx', 'serve')
